"""Leads import service — parse CSV/XLSX files and map columns to Lead fields."""

from __future__ import annotations

import csv
import io
import json
import logging
from typing import Optional

logger = logging.getLogger(__name__)

# ── Column mapping hints ──────────────────────────────────
# Maps common CSV header names (lowercased) to our Lead field names.

COLUMN_ALIASES: dict[str, str] = {
    # First name
    "first_name": "first_name",
    "first name": "first_name",
    "firstname": "first_name",
    "fname": "first_name",
    "owner first name": "first_name",
    "owner_first_name": "first_name",
    # Last name
    "last_name": "last_name",
    "last name": "last_name",
    "lastname": "last_name",
    "lname": "last_name",
    "owner last name": "last_name",
    "owner_last_name": "last_name",
    # Full name
    "full_name": "full_name",
    "full name": "full_name",
    "name": "full_name",
    "owner name": "full_name",
    "owner_name": "full_name",
    "contact name": "full_name",
    # Phone
    "phone": "phone",
    "phone_number": "phone",
    "phone number": "phone",
    "telephone": "phone",
    "mobile": "phone",
    "cell": "phone",
    "owner phone": "phone",
    # Email
    "email": "email",
    "email_address": "email",
    "email address": "email",
    "e-mail": "email",
    "owner email": "email",
    # Address
    "address": "address",
    "street": "address",
    "street_address": "address",
    "street address": "address",
    "property address": "address",
    "property_address": "address",
    "mailing address": "address",
    "mailing_address": "address",
    # City
    "city": "city",
    "property city": "city",
    "property_city": "city",
    "mailing city": "city",
    # State
    "state": "state",
    "st": "state",
    "property state": "state",
    "property_state": "state",
    "mailing state": "state",
    # Zip
    "zip": "zip_code",
    "zip_code": "zip_code",
    "zip code": "zip_code",
    "zipcode": "zip_code",
    "postal": "zip_code",
    "postal_code": "zip_code",
    "postal code": "zip_code",
    "property zip": "zip_code",
    "mailing zip": "zip_code",
    # Property type
    "property_type": "property_type",
    "property type": "property_type",
    "type": "property_type",
    "prop type": "property_type",
    "prop_type": "property_type",
}

VALID_FIELDS = {
    "first_name", "last_name", "full_name", "phone", "email",
    "address", "city", "state", "zip_code", "property_type",
}


def detect_column_mapping(headers: list[str]) -> dict[str, str]:
    """Auto-detect column mappings from CSV/XLSX headers.

    Returns dict of { original_header: our_field_name } for detected matches.
    Unmatched headers are included with value "" (unmapped).
    """
    mapping: dict[str, str] = {}
    used_fields: set[str] = set()

    for header in headers:
        normalized = header.strip().lower()
        if normalized in COLUMN_ALIASES:
            field = COLUMN_ALIASES[normalized]
            if field not in used_fields:
                mapping[header] = field
                used_fields.add(field)
            else:
                mapping[header] = ""  # Duplicate mapping — leave unmapped
        else:
            mapping[header] = ""  # No match found

    return mapping


def parse_csv_content(content: bytes, encoding: str = "utf-8") -> tuple[list[str], list[dict]]:
    """Parse CSV content into headers and rows.

    Returns: (headers, rows) where rows are list of dicts keyed by header names.
    """
    text = content.decode(encoding, errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)
    return headers, rows


def parse_xlsx_content(content: bytes) -> tuple[list[str], list[dict]]:
    """Parse XLSX content into headers and rows.

    Returns: (headers, rows) where rows are list of dicts keyed by header names.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []

    rows_iter = ws.iter_rows(values_only=True)

    # First row is headers
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(raw_headers)]

    rows = []
    for row_values in rows_iter:
        if all(v is None for v in row_values):
            continue  # Skip empty rows
        row_dict = {}
        for i, val in enumerate(row_values):
            if i < len(headers):
                row_dict[headers[i]] = str(val).strip() if val is not None else ""
        rows.append(row_dict)

    wb.close()
    return headers, rows


def apply_mapping(rows: list[dict], mapping: dict[str, str]) -> list[dict]:
    """Apply column mapping to transform raw rows into Lead-field rows.

    Only includes fields that are mapped (non-empty mapping values).
    Also computes full_name from first_name + last_name if full_name not directly mapped.
    """
    mapped_rows = []
    for raw_row in rows:
        lead_data: dict[str, str] = {}
        for original_header, our_field in mapping.items():
            if our_field and our_field in VALID_FIELDS:
                value = raw_row.get(original_header, "").strip()
                if value:
                    lead_data[our_field] = value

        # Compute full_name if not directly mapped
        if not lead_data.get("full_name"):
            first = lead_data.get("first_name", "")
            last = lead_data.get("last_name", "")
            if first or last:
                lead_data["full_name"] = f"{first} {last}".strip()

        # Only include rows that have at least a name or address
        has_name = lead_data.get("full_name") or lead_data.get("first_name") or lead_data.get("last_name")
        has_address = lead_data.get("address")
        if has_name or has_address:
            mapped_rows.append(lead_data)

    return mapped_rows
